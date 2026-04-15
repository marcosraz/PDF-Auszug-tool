#!/usr/bin/env python3
"""
Improved PDF Extractor for Kujira Piping Isometrics
Uses Claude Opus 4.5 Vision for accurate extraction
"""

import os
import sys
import json
import base64
import subprocess
import tempfile
import re
from pathlib import Path
from datetime import datetime

import requests
from openpyxl import Workbook
from openpyxl.styles import Font
from PIL import Image

# Expected output columns (matching the template format)
OUTPUT_COLUMNS = [
    "Rev.", "Latest Rev.", "Length", "Name", "PId", "Pipe class", "PRIOs", "HOLD",
    "Building", "Floor", "Insulation", "Transmittal No.", "Date Rec.", "Spooling",
    "Iss. to prefab", "Prefab remarks", "Prefab start", "Welded", "Latest location",
    "Sent to site", "Work package", "Prefab compl.", "Install start", "Install 50",
    "Install 90", "Install compl.", "Area", "Company", "Site remarks", "As built",
    "Site checked", "Prefab Company", "Ped Cat", "Ped Cat.1", "Rel. for Insulation",
    "Insulation Start", "Insulation 50%", "Insulation 90%", "Insulated", "Insulated.1",
    "Cladding", "Cladding.1"
]

# Fields we can extract from PDFs
EXTRACTABLE_FIELDS = [
    "Rev.", "Length", "Name", "PId", "Pipe class", "DN",
    "Building", "Floor", "Insulation", "Ped Cat"
]

EXTRACTION_PROMPT = """Du bist ein Experte für Piping Isometric Zeichnungen von AMELICOR/Lonza.

Analysiere diese Zeichnung und extrahiere die Daten aus dem TITELBLOCK (unten rechts) und der MATERIALLISTE (oben rechts).

KRITISCHE REGELN FÜR LINE NUMBER FORMAT:
Das Format ist: 740-[TYPE]-[VESSEL]-[XXX]-[PIPECLASS]-[DN]-[SUFFIX]

Beispiele:
- "740-ABkont-VP131-102-LH099-20-0"
  → Type: ABkont, Vessel: VP131, XXX: 102, Pipe class: LH099, DN: 20
- "740-AB-VU200-201-LH032-25-0"
  → Type: AB, Vessel: VU200, XXX: 201, Pipe class: LH032, DN: 25
- "740-CIP-S-VP130-102-LH089-25-0"
  → Type: CIP-S, Vessel: VP130, XXX: 102, Pipe class: LH089, DN: 25

WICHTIG:
- Line No./Name: Die EXAKTE Zeichnungsnummer aus dem Titelblock (z.B. "740-ABkont-VP131-102-LH099-20-0")
- NICHT den Dateinamen verwenden!
- Rev.: Revisionsnummer als Integer (z.B. 1, 2)
- Length: SUMME aller Rohrlängen aus der Materialliste unter "ROHRE" (z.B. 16.5M + 0.5M = 17.0)
- PId: Die P&ID Nummer (z.B. "LV1-105638.VU311")
- Pipe class: Der PIPECLASS-Code aus der Line Number (z.B. "LH099", "LH032", "LH089") - NICHT "102"!
- DN: Der DN-Wert aus der Line Number (vorletzte Zahl, z.B. 20, 25) - NICHT "102"!
- Building: Kundenname ist "Lonza AG" - schreibe "MC1" wenn nicht anders angegeben
- Floor: Level/Etage (z.B. "Level 5", "Level 6")
- Insulation: Falls vorhanden
- Ped Cat: PED Kategorie falls vorhanden

Antworte NUR mit JSON (keine Erklärungen):
{
  "line_no": "740-ABkont-VP131-102-LH099-20-0",
  "rev": 1,
  "length": 1.0,
  "pid": "LV1-105638.VU311",
  "pipe_class": "LH099",
  "dn": 20,
  "building": "MC1",
  "floor": "Level 5",
  "insulation": null,
  "ped_cat": null
}"""


def pdf_to_image(pdf_path: Path, output_dir: Path, dpi: int = 200) -> list[Path]:
    """Convert PDF pages to PNG images using pdftoppm"""
    output_prefix = output_dir / pdf_path.stem

    cmd = [
        "pdftoppm",
        "-png",
        "-r", str(dpi),
        str(pdf_path),
        str(output_prefix)
    ]

    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise RuntimeError(f"pdftoppm failed: {result.stderr}")

    images = sorted(output_dir.glob(f"{pdf_path.stem}*.png"))
    return images


def image_to_base64(image_path: Path, max_size: int = 2048) -> str:
    """Convert image to base64, resizing if necessary"""
    with Image.open(image_path) as img:
        if max(img.size) > max_size:
            ratio = max_size / max(img.size)
            new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
            img = img.resize(new_size, Image.Resampling.LANCZOS)

        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        import io
        buffer = io.BytesIO()
        img.save(buffer, format='PNG', optimize=True)
        return base64.standard_b64encode(buffer.getvalue()).decode('utf-8')


def parse_json_response(text: str) -> dict:
    """Parse JSON from response"""
    # Remove markdown code blocks
    if "```json" in text:
        match = re.search(r'```json\s*([\s\S]*?)```', text)
        if match:
            text = match.group(1).strip()
    elif "```" in text:
        match = re.search(r'```\s*([\s\S]*?)```', text)
        if match:
            text = match.group(1).strip()

    try:
        return json.loads(text)
    except (json.JSONDecodeError, ValueError):
        pass

    # Fallback: find JSON object
    match = re.search(r'\{[\s\S]*\}', text)
    if match:
        try:
            return json.loads(match.group())
        except (json.JSONDecodeError, ValueError):
            pass

    return {}


def extract_with_claude_opus(image_base64: str, api_key: str) -> dict:
    """Use Claude Opus 4.5 Vision API for extraction"""
    url = "https://api.anthropic.com/v1/messages"

    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01"
    }

    payload = {
        "model": "claude-opus-4-5-20250514",
        "max_tokens": 2048,
        "messages": [{
            "role": "user",
            "content": [
                {
                    "type": "image",
                    "source": {
                        "type": "base64",
                        "media_type": "image/png",
                        "data": image_base64
                    }
                },
                {"type": "text", "text": EXTRACTION_PROMPT}
            ]
        }]
    }

    response = requests.post(url, headers=headers, json=payload, timeout=120)

    if response.status_code != 200:
        raise RuntimeError(f"Claude API error: {response.status_code} - {response.text}")

    result = response.json()
    text_content = result["content"][0]["text"]

    return parse_json_response(text_content)


def validate_and_fix_extraction(data: dict, filename: str) -> dict:
    """Validate and fix common extraction errors"""
    line_no = data.get('line_no', '')

    # Parse line number to validate/fix pipe_class and dn
    # Format: 740-[TYPE]-[VESSEL]-[XXX]-[PIPECLASS]-[DN]-[SUFFIX]
    parts = line_no.split('-')

    if len(parts) >= 6:
        # Find pipe class (starts with LH, I, SP, etc.)
        for i, part in enumerate(parts):
            if re.match(r'^(LH|I\d|SP)\d*', part):
                data['pipe_class'] = part
                # DN is usually the next part
                if i + 1 < len(parts) and parts[i + 1].isdigit():
                    data['dn'] = int(parts[i + 1])
                break

    # Validate pipe_class isn't just a number
    pipe_class = data.get('pipe_class', '')
    if pipe_class and pipe_class.isdigit():
        # Try to find correct pipe class from line_no
        for part in parts:
            if re.match(r'^(LH|I)\d+', part):
                data['pipe_class'] = part
                break

    # Validate DN is reasonable (typically 15-450)
    dn = data.get('dn')
    if dn and (dn > 500 or dn < 10):
        # Try to extract from line_no
        match = re.search(r'-(\d{2,3})-\d$', line_no)
        if match:
            data['dn'] = int(match.group(1))

    return data


def process_pdf(pdf_path: Path, api_key: str) -> dict:
    """Process a single PDF and extract data"""
    print(f"Processing: {pdf_path.name}")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        images = pdf_to_image(pdf_path, tmpdir)

        if not images:
            print(f"  -> No images generated", file=sys.stderr)
            return {}

        image_b64 = image_to_base64(images[0])

        try:
            data = extract_with_claude_opus(image_b64, api_key)
            data = validate_and_fix_extraction(data, pdf_path.name)
            print(f"  -> Extracted: {data.get('line_no', 'N/A')}")
            return data
        except Exception as e:
            print(f"  -> Error: {e}", file=sys.stderr)
            return {}


def create_excel(data_list: list[dict], output_path: Path):
    """Create Excel file matching the expected template format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Extracted Data"

    # Write headers
    for col, header in enumerate(OUTPUT_COLUMNS, 1):
        ws.cell(row=1, column=col, value=header)
        ws.cell(row=1, column=col).font = Font(bold=True)

    # Write data
    for row_idx, data in enumerate(data_list, 2):
        line_no = data.get('line_no', '')

        # Map extracted data to output columns
        ws.cell(row=row_idx, column=1, value=data.get('rev'))  # Rev.
        ws.cell(row=row_idx, column=2, value=True)  # Latest Rev.
        ws.cell(row=row_idx, column=3, value=data.get('length'))  # Length
        ws.cell(row=row_idx, column=4, value=line_no)  # Name
        ws.cell(row=row_idx, column=5, value=data.get('pid'))  # PId
        ws.cell(row=row_idx, column=6, value=data.get('pipe_class'))  # Pipe class
        # PRIOs, HOLD - empty (columns 7-8)
        ws.cell(row=row_idx, column=9, value=data.get('building'))  # Building
        ws.cell(row=row_idx, column=10, value=data.get('floor'))  # Floor
        ws.cell(row=row_idx, column=11, value=data.get('insulation'))  # Insulation
        # Transmittal No. through Cladding.1 - mostly empty workflow fields
        ws.cell(row=row_idx, column=33, value=data.get('ped_cat'))  # Ped Cat
        ws.cell(row=row_idx, column=34, value=data.get('ped_cat'))  # Ped Cat.1

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except (TypeError, ValueError):
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 40)

    wb.save(output_path)
    print(f"\nExcel saved to: {output_path}")


def main():
    """Main function"""
    api_key = os.environ.get('ANTHROPIC_API_KEY')
    if not api_key:
        print("Error: Set ANTHROPIC_API_KEY environment variable", file=sys.stderr)
        sys.exit(1)

    # Kujira folder
    kujira_dir = Path(__file__).parent.parent / "Kujira"

    if not kujira_dir.exists():
        print(f"Error: Kujira directory not found at {kujira_dir}", file=sys.stderr)
        sys.exit(1)

    pdf_files = sorted(kujira_dir.glob("*.pdf"))
    print(f"Found {len(pdf_files)} PDF files in {kujira_dir}")

    # Process all PDFs
    all_data = []
    for pdf in pdf_files:
        data = process_pdf(pdf, api_key)
        if data:
            data['_source_file'] = pdf.name
            all_data.append(data)

    # Create output Excel
    output_path = Path(__file__).parent / "Kujira_IMPROVED.xlsx"
    create_excel(all_data, output_path)

    print(f"\nProcessed {len(all_data)}/{len(pdf_files)} PDFs successfully")

    # Print summary
    print("\n=== EXTRACTION SUMMARY ===")
    for data in all_data:
        print(f"  {data.get('line_no', 'N/A')}: Rev={data.get('rev')}, "
              f"DN={data.get('dn')}, Pipe={data.get('pipe_class')}, "
              f"Length={data.get('length')}m")


if __name__ == "__main__":
    main()
