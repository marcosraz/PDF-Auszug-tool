#!/usr/bin/env python3
"""
PDF Data Extractor - Extracts piping isometric data from PDFs to Excel
Supports Claude Vision and Gemini with Few-Shot Learning
"""

import os
import sys
import json
import base64
import subprocess
import tempfile
import re
import random
from pathlib import Path
from typing import Optional
from datetime import datetime

import click
import requests
from openpyxl import Workbook, load_workbook
from PIL import Image
import google.auth.transport.requests
from google.oauth2 import service_account
from tenacity import retry, wait_exponential_jitter, stop_after_attempt, retry_if_exception_type


# Paths
SCRIPT_DIR = Path(__file__).parent
EXAMPLES_DIR = SCRIPT_DIR / "examples"
SERVICE_ACCOUNT_JSON = SCRIPT_DIR / "zg-visual-recognition-65464-f332a04f02a3.json"
VERTEX_AI_PROJECT = "zg-visual-recognition-65464"
VERTEX_AI_LOCATION = "global"
DEFAULT_GEMINI_MODEL = "gemini-3.1-pro-preview"

# Full 44-column template matching customer format
EXCEL_COLUMNS = [
    "Id", "Line No.", "Rev.", "Latest Rev.", "Length", "Name", "PId",
    "Pipe class", "PRIOs", "HOLD", "Building", "Floor", "Insulation",
    "Transmittal No.", "Date Rec.", "Spooling", "Iss. to prefab",
    "Prefab remarks", "Prefab start", "Welded", "Latest location",
    "Sent to site", "Work package", "Prefab compl.", "Install start",
    "Install 50", "Install 90", "Install compl.", "Area", "Company",
    "Site remarks", "As built", "Site checked", "Prefab Company",
    "Ped Cat", "Ped Cat", "Rel. for Insulation", "Insulation Start",
    "Insulation 50%", "Insulation 90%", "Insulated", "Insulated",
    "Cladding", "Cladding"
]

EXTRACTION_PROMPT = """Du bist ein Experte für Piping Isometric Zeichnungen. Extrahiere NUR die SICHTBAREN Daten aus dem Titelblock.

WICHTIG - Lies genau vom Bild ab, erfinde NICHTS:

1. line_no: Die Isometric-Nummer / Zeichnungsnummer (z.B. "740-ABkont-VP131-102-LH099-20-0", "CSL-7-201-017-25-SP1")
   - Steht meist unten rechts als "Isometric Number" oder große Beschriftung

2. rev: Revisionsnummer (0, 1, 2) - aus Revisionsfeld oder Dateiname

3. length: Gesamtlänge in Metern - aus "CUT PIPE LENGTH" Tabelle oder Materialliste summieren
   - Achte auf die Einheit (m oder mm)!

4. pid: P&ID Nummer (z.B. "LV1-105638.VU311", "CH06822004-BD-PI-1179")
   - Steht als "P&ID-No." oder "Tag/P&ID" im Titelblock

5. pipe_class: NUR wenn ein Feld "Pipe class" oder "Pipe spec" SICHTBAR ist (z.B. "SP1", "SD", "LH099")
   - NICHT aus der Line No. raten! Wenn nicht sichtbar: null

6. building: NUR wenn "Building" oder "Gebäude" SICHTBAR ist (z.B. "MC1", "DC building")
   - ACHTUNG: "Lonza AG", "ThermoFisher" sind KUNDEN, nicht Buildings!
   - Wenn nicht sichtbar: null

7. floor: Etage/Level (z.B. "Level 5", "LVL 6", "F02")
   - Steht oft als "Floor" oder in einem separaten Feld

8. dn: Nennweite in mm - aus Materialliste (DN-Spalte) oder aus Line No. (z.B. -25- = DN25)

9. insulation: Isolierungstyp (z.B. "W60", "H2", "K51", "N/A")
   - Steht als "Insulation" im Titelblock

10. project: Projektname (z.B. "Kujira", "5K SUT", "LPP 6A")

11. ped_cat: PED Kategorie - NUR wenn sichtbar (z.B. "SEP")

12. customer: Kundenname (z.B. "Lonza AG", "ThermoFisher")

Antworte NUR mit JSON. Für JEDES Feld gib einen Confidence-Score (0.0-1.0) an, der angibt wie sicher du dir bist:
{"line_no": {"value": "...", "confidence": 0.95}, "rev": {"value": 0, "confidence": 0.9}, "length": {"value": 0.0, "confidence": 0.7}, "pid": {"value": "...", "confidence": 0.85}, "pipe_class": {"value": null, "confidence": 0.8}, "building": {"value": null, "confidence": 0.9}, "floor": {"value": "...", "confidence": 0.8}, "dn": {"value": 0, "confidence": 0.85}, "insulation": {"value": null, "confidence": 0.8}, "project": {"value": "...", "confidence": 0.9}, "ped_cat": {"value": null, "confidence": 0.8}, "customer": {"value": "...", "confidence": 0.9}}
"""


def load_examples(project: str = None) -> list[dict]:
    """Load few-shot examples from examples directory.

    Args:
        project: Optional project name filter (kept for backward compat, but
                 filtering is now handled by select_examples()).
    """
    examples = []

    if not EXAMPLES_DIR.exists():
        return examples

    for json_file in sorted(EXAMPLES_DIR.glob("*.json")):
        image_file = json_file.with_suffix(".png")
        if not image_file.exists():
            continue

        with open(image_file, "rb") as f:
            image_b64 = base64.standard_b64encode(f.read()).decode('utf-8')

        with open(json_file) as f:
            correct_output = json.load(f)

        examples.append({
            "name": json_file.stem,
            "image_b64": image_b64,
            "output": correct_output
        })

    return examples


def select_examples(all_examples: list[dict], project: str = None,
                    max_examples: int = 5) -> list[dict]:
    """Smart few-shot example selection prioritizing project-specific examples.

    Args:
        all_examples: All loaded examples.
        project: Optional project name to prioritize matching examples.
        max_examples: Maximum number of examples to return (default 5).

    Returns:
        Selected subset of examples, prioritizing same-project examples
        while keeping at least 1 from a different project for diversity.
    """
    if not all_examples:
        return []

    if len(all_examples) <= max_examples:
        return all_examples

    if not project:
        # No project hint - return random selection
        return random.sample(all_examples, min(max_examples, len(all_examples)))

    project_lower = project.lower()

    # Split into same-project and other-project examples
    same_project = []
    other_project = []
    for ex in all_examples:
        ex_project = (ex.get("output", {}).get("project") or "").lower()
        ex_name = ex.get("name", "").lower()
        if project_lower in ex_project or project_lower in ex_name:
            same_project.append(ex)
        else:
            other_project.append(ex)

    # If no project-specific examples, fall back to random selection
    if not same_project:
        return random.sample(all_examples, min(max_examples, len(all_examples)))

    selected = []

    # Reserve at least 1 slot for diversity (different project)
    diversity_count = min(1, len(other_project))
    same_slots = max_examples - diversity_count

    # Add same-project examples (up to same_slots)
    if len(same_project) <= same_slots:
        selected.extend(same_project)
    else:
        selected.extend(random.sample(same_project, same_slots))

    # Fill remaining slots with other-project examples for diversity
    remaining = max_examples - len(selected)
    if remaining > 0 and other_project:
        selected.extend(random.sample(other_project, min(remaining, len(other_project))))

    return selected


def pdf_to_image(pdf_path: Path, output_dir: Path, dpi: int = 150) -> list[Path]:
    """Convert PDF pages to PNG images using PyMuPDF"""
    import fitz

    doc = fitz.open(str(pdf_path))
    images = []
    zoom = dpi / 72
    matrix = fitz.Matrix(zoom, zoom)

    for page_num in range(len(doc)):
        page = doc[page_num]
        pix = page.get_pixmap(matrix=matrix)
        output_path = output_dir / f"{pdf_path.stem}-{page_num + 1:02d}.png"
        pix.save(str(output_path))
        images.append(output_path)

    doc.close()
    return images


def image_to_base64(image_path: Path, max_size: int = 1568) -> str:
    """Convert image to base64, resizing if necessary for API limits"""
    with Image.open(image_path) as img:
        if max(img.size) > max_size:
            ratio = max_size / max(img.size)
            new_size = (int(img.size[0] * ratio), int(img.size[1] * ratio))
            img = img.resize(new_size, Image.Resampling.LANCZOS)

        if img.mode in ('RGBA', 'P'):
            img = img.convert('RGB')

        import io
        buffer = io.BytesIO()
        img.save(buffer, format='PNG')
        return base64.standard_b64encode(buffer.getvalue()).decode('utf-8')


def parse_json_response(text: str) -> dict:
    """Parse JSON from various response formats"""
    # Remove markdown code blocks
    if "```json" in text:
        match = re.search(r'```json\s*([\s\S]*?)```', text)
        if match:
            text = match.group(1).strip()
    elif "```" in text:
        match = re.search(r'```\s*([\s\S]*?)```', text)
        if match:
            text = match.group(1).strip()

    # Try parsing cleaned text
    try:
        return json.loads(text)
    except:
        pass

    # Fallback: find JSON object (supports nested braces for confidence format)
    # First try nested match, then flat
    match = re.search(r'\{(?:[^{}]|\{[^{}]*\})*\}', text, re.DOTALL)
    if match:
        try:
            return json.loads(match.group())
        except:
            pass

    return {}


def parse_extraction_with_confidence(text: str) -> tuple[dict, dict]:
    """Parse JSON response that may contain confidence scores.

    Returns (data, confidence) where:
    - data: dict of field -> value (backward compatible)
    - confidence: dict of field -> float (0.0-1.0)

    Handles both new format {"field": {"value": ..., "confidence": 0.9}}
    and old format {"field": value} gracefully.
    """
    raw = parse_json_response(text)
    if not raw:
        return {}, {}

    data = {}
    confidence = {}

    for key, val in raw.items():
        if isinstance(val, dict) and "value" in val:
            # New format with confidence
            data[key] = val["value"]
            confidence[key] = float(val.get("confidence", 0.0))
        else:
            # Old format without confidence - fall back gracefully
            data[key] = val

    return data, confidence


def extract_with_claude(image_base64: str, api_key: str) -> dict:
    """Use Claude Vision API to extract data from image"""
    url = "https://api.anthropic.com/v1/messages"

    headers = {
        "Content-Type": "application/json",
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01"
    }

    payload = {
        "model": "claude-sonnet-4-20250514",
        "max_tokens": 1024,
        "messages": [{
            "role": "user",
            "content": [
                {"type": "image", "source": {"type": "base64", "media_type": "image/png", "data": image_base64}},
                {"type": "text", "text": EXTRACTION_PROMPT}
            ]
        }]
    }

    response = requests.post(url, headers=headers, json=payload, timeout=60)

    if response.status_code != 200:
        raise RuntimeError(f"Claude API error: {response.status_code} - {response.text}")

    result = response.json()
    text_content = result["content"][0]["text"]

    return parse_json_response(text_content)


def get_vertex_ai_token(credentials_path: Path = None) -> str:
    """Get access token from service account JSON for Vertex AI"""
    creds_path = credentials_path or SERVICE_ACCOUNT_JSON
    creds = service_account.Credentials.from_service_account_file(
        str(creds_path),
        scopes=["https://www.googleapis.com/auth/cloud-platform"]
    )
    creds.refresh(google.auth.transport.requests.Request())
    return creds.token


def extract_with_gemini(image_base64: str, api_key: str = None, model: str = None,
                        examples: list[dict] = None, credentials_path: Path = None,
                        project: str = None) -> tuple[dict, dict]:
    """Use Gemini Vision API via Vertex AI (service account) or API key fallback.

    Returns (data, confidence) tuple. data contains field values,
    confidence contains field -> float scores.
    """

    model = model or DEFAULT_GEMINI_MODEL

    # Build contents with few-shot examples
    contents = []

    if examples:
        for i, ex in enumerate(examples):
            contents.append({
                "role": "user",
                "parts": [
                    {"inline_data": {"mime_type": "image/png", "data": ex["image_b64"]}},
                    {"text": f"Beispiel {i+1}: Extrahiere die Daten aus diesem Piping Isometric."}
                ]
            })
            contents.append({
                "role": "model",
                "parts": [{"text": json.dumps(ex["output"], indent=2)}]
            })

    # Add actual image (system prompt is in systemInstruction for cacheability)
    contents.append({
        "role": "user",
        "parts": [
            {"inline_data": {"mime_type": "image/png", "data": image_base64}},
            {"text": "Extrahiere jetzt die Daten aus diesem Piping Isometric:"}
        ]
    })

    # Structured output schema for Gemini - each field has value + confidence
    _field_schema = {
        "type": "OBJECT",
        "properties": {
            "value": {"type": "STRING", "nullable": True},
            "confidence": {"type": "NUMBER"},
        },
        "required": ["value", "confidence"],
    }
    _response_schema = {
        "type": "OBJECT",
        "properties": {
            "line_no": _field_schema,
            "rev": _field_schema,
            "length": _field_schema,
            "pid": _field_schema,
            "pipe_class": _field_schema,
            "building": _field_schema,
            "floor": _field_schema,
            "dn": _field_schema,
            "insulation": _field_schema,
            "project": _field_schema,
            "ped_cat": _field_schema,
            "customer": _field_schema,
        },
        "required": [
            "line_no", "rev", "length", "pid", "pipe_class",
            "building", "floor", "dn", "insulation", "project",
            "ped_cat", "customer",
        ],
    }

    payload = {
        "contents": contents,
        "systemInstruction": {
            "parts": [{"text": EXTRACTION_PROMPT}]
        },
        "generationConfig": {
            "temperature": 0.1,
            "maxOutputTokens": 8192,
            "responseMimeType": "application/json",
            "responseSchema": _response_schema,
        },
    }

    timeout = 120

    # Prefer service account via Vertex AI
    creds_path = credentials_path or SERVICE_ACCOUNT_JSON
    if creds_path.exists():
        token = get_vertex_ai_token(creds_path)
        if VERTEX_AI_LOCATION == "global":
            url = (
                f"https://aiplatform.googleapis.com/v1/"
                f"projects/{VERTEX_AI_PROJECT}/locations/{VERTEX_AI_LOCATION}/"
                f"publishers/google/models/{model}:generateContent"
            )
        else:
            url = (
                f"https://{VERTEX_AI_LOCATION}-aiplatform.googleapis.com/v1/"
                f"projects/{VERTEX_AI_PROJECT}/locations/{VERTEX_AI_LOCATION}/"
                f"publishers/google/models/{model}:generateContent"
            )
        headers = {
            "Content-Type": "application/json",
            "Authorization": f"Bearer {token}"
        }
    elif api_key:
        # Fallback to API key
        url = f"https://generativelanguage.googleapis.com/v1beta/models/{model}:generateContent?key={api_key}"
        headers = {"Content-Type": "application/json"}
    else:
        raise RuntimeError("No credentials found. Place service account JSON or set GEMINI_API_KEY.")

    @retry(
        wait=wait_exponential_jitter(initial=2, max=60, jitter=5),
        stop=stop_after_attempt(4),
        retry=retry_if_exception_type((requests.exceptions.Timeout, requests.exceptions.ConnectionError)),
        reraise=True,
    )
    def _call_api():
        resp = requests.post(url, headers=headers, json=payload, timeout=timeout)
        if resp.status_code == 429 or resp.status_code >= 500:
            raise requests.exceptions.ConnectionError(
                f"Retryable error: {resp.status_code} - {resp.text[:200]}"
            )
        return resp

    try:
        response = _call_api()
    except requests.exceptions.Timeout:
        raise RuntimeError(f"Timeout after {timeout}s and retries exhausted")
    except requests.exceptions.ConnectionError as e:
        raise RuntimeError(f"Gemini API unreachable after retries: {e}")

    if response.status_code != 200:
        raise RuntimeError(f"Gemini API error: {response.status_code} - {response.text[:200]}")

    result = response.json()

    try:
        text = result["candidates"][0]["content"]["parts"][0]["text"]
    except (KeyError, IndexError) as e:
        raise RuntimeError(f"Response parse error: {e}")

    return parse_extraction_with_confidence(text)


def process_pdf(pdf_path: Path, api_key: str = None, provider: str = "gemini",
                model: str = None, use_fewshot: bool = True,
                project: str = None) -> tuple[dict, dict]:
    """Process a single PDF and extract data.

    Returns (data, confidence) tuple.
    """
    click.echo(f"Processing: {pdf_path.name}")

    # Load examples for few-shot with smart selection
    if use_fewshot:
        all_examples = load_examples()
        examples = select_examples(all_examples, project=project)
    else:
        examples = []
    if examples and use_fewshot:
        click.echo(f"  Using {len(examples)} few-shot examples")

    with tempfile.TemporaryDirectory() as tmpdir:
        tmpdir = Path(tmpdir)
        images = pdf_to_image(pdf_path, tmpdir)

        if not images:
            click.echo(f"  -> No images generated", err=True)
            return {}, {}

        image_b64 = image_to_base64(images[0])

        try:
            if provider == "claude":
                data = extract_with_claude(image_b64, api_key)
                click.echo(f"  -> Extracted via Claude Vision")
                return data, {}
            else:
                model = model or DEFAULT_GEMINI_MODEL
                data, confidence = extract_with_gemini(
                    image_b64, api_key=api_key, model=model,
                    examples=examples, project=project
                )
                fewshot_info = " (few-shot)" if examples else ""
                click.echo(f"  -> Extracted via {model}{fewshot_info}")

            return data, confidence
        except Exception as e:
            click.echo(f"  -> Error: {e}", err=True)
            return {}, {}


def create_excel(data_list: list[dict], output_path: Path, template_path: Optional[Path] = None):
    """Create Excel file from extracted data using 44-column customer format"""
    if template_path and template_path.exists():
        wb = load_workbook(template_path)
        ws = wb.active
        start_row = ws.max_row + 1
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Extracted Data"

        # Use full 44-column format
        for col, header in enumerate(EXCEL_COLUMNS, 1):
            ws.cell(row=1, column=col, value=header)

        from openpyxl.styles import Font
        for cell in ws[1]:
            cell.font = Font(bold=True)

        start_row = 2

    for row_idx, data in enumerate(data_list, start_row):
        line_no = data.get('line_no')

        # Column mapping to 44-column format:
        # 1=Id, 2=Line No., 3=Rev., 4=Latest Rev., 5=Length, 6=Name, 7=PId
        # 8=Pipe class, 9=PRIOs, 10=HOLD, 11=Building, 12=Floor, 13=Insulation
        # 14=Transmittal No., 15=Date Rec., ... 35=Ped Cat, 36=Ped Cat, ...

        ws.cell(row=row_idx, column=1, value=row_idx - 1)  # Id
        ws.cell(row=row_idx, column=2, value=line_no)  # Line No.
        ws.cell(row=row_idx, column=3, value=data.get('rev'))  # Rev.
        ws.cell(row=row_idx, column=4, value='false')  # Latest Rev. - needs manual check
        ws.cell(row=row_idx, column=5, value=data.get('length'))  # Length
        ws.cell(row=row_idx, column=6, value=line_no)  # Name (same as Line No.)
        ws.cell(row=row_idx, column=7, value=data.get('pid'))  # PId
        ws.cell(row=row_idx, column=8, value=data.get('pipe_class'))  # Pipe class
        # 9=PRIOs, 10=HOLD - leave empty
        ws.cell(row=row_idx, column=11, value=data.get('building'))  # Building
        ws.cell(row=row_idx, column=12, value=data.get('floor'))  # Floor
        ws.cell(row=row_idx, column=13, value=data.get('insulation') or 'N/A')  # Insulation
        # 14-28 = Workflow fields - leave empty
        # 29=Area, 30=Company - leave empty (internal IDs)
        # 31-34 = More workflow - leave empty
        ws.cell(row=row_idx, column=35, value=data.get('ped_cat'))  # Ped Cat
        ws.cell(row=row_idx, column=36, value=data.get('ped_cat'))  # Ped Cat (duplicate)
        # 37-44 = Insulation workflow - leave empty

    # Auto-adjust widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)

    wb.save(output_path)
    try:
        click.echo(f"\nExcel saved to: {output_path}")
    except RuntimeError:
        pass  # Suppress click.echo errors when called outside CLI context


@click.group()
@click.version_option(version="2.0.0")
def cli():
    """PDF Data Extractor - Extract piping isometric data from PDFs to Excel

    Now with Gemini 3 Flash + Few-Shot Learning for better accuracy!
    """
    pass


@cli.command()
@click.argument('input_path', type=click.Path(exists=True))
@click.option('--output', '-o', type=click.Path(), help='Output Excel file path')
@click.option('--provider', type=click.Choice(['gemini', 'claude']), default='gemini',
              help='AI provider (default: gemini)')
@click.option('--model', '-m', type=str, help=f'Model name (default: {DEFAULT_GEMINI_MODEL})')
@click.option('--no-fewshot', is_flag=True, help='Disable few-shot learning')
@click.option('--recursive', '-r', is_flag=True, help='Process subdirectories recursively')
def extract(input_path: str, output: Optional[str], provider: str, model: str,
            no_fewshot: bool, recursive: bool):
    """Extract data from PDF(s) and create Excel file

    INPUT_PATH can be a single PDF file or a directory containing PDFs.
    """
    # Get API key based on provider
    if provider == "claude":
        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            click.echo("Error: Set ANTHROPIC_API_KEY environment variable", err=True)
            sys.exit(1)
    else:
        api_key = os.environ.get('GEMINI_API_KEY')
        if not api_key and not SERVICE_ACCOUNT_JSON.exists():
            click.echo("Error: Set GEMINI_API_KEY or place service account JSON", err=True)
            sys.exit(1)
        if SERVICE_ACCOUNT_JSON.exists():
            click.echo(f"Using service account: {SERVICE_ACCOUNT_JSON.name}")

    input_path = Path(input_path)

    # Collect PDFs
    if input_path.is_file():
        if input_path.suffix.lower() != '.pdf':
            click.echo("Error: Input file must be a PDF", err=True)
            sys.exit(1)
        pdf_files = [input_path]
        default_output = input_path.parent / f"{input_path.stem}_extracted.xlsx"
    else:
        pattern = '**/*.pdf' if recursive else '*.pdf'
        pdf_files = sorted(input_path.glob(pattern))
        default_output = input_path / "extracted_data.xlsx"

    if not pdf_files:
        click.echo("No PDF files found", err=True)
        sys.exit(1)

    click.echo(f"Found {len(pdf_files)} PDF file(s)")
    click.echo(f"Provider: {provider} | Few-Shot: {'OFF' if no_fewshot else 'ON'}")

    # Process each PDF
    all_data = []
    for pdf in pdf_files:
        data, confidence = process_pdf(pdf, api_key, provider=provider, model=model,
                                       use_fewshot=not no_fewshot)
        if data:
            data['_source_file'] = pdf.name
            if confidence:
                data['_confidence'] = confidence
            all_data.append(data)

    # Create Excel
    output_path = Path(output) if output else default_output
    create_excel(all_data, output_path)

    click.echo(f"\nProcessed {len(all_data)}/{len(pdf_files)} PDFs successfully")


@cli.command()
@click.argument('pdf_path', type=click.Path(exists=True))
@click.option('--provider', type=click.Choice(['gemini', 'claude']), default='gemini')
@click.option('--model', '-m', type=str, default=None)
@click.option('--no-fewshot', is_flag=True, help='Disable few-shot learning')
def test(pdf_path: str, provider: str, model: str, no_fewshot: bool):
    """Test extraction on a single PDF and show results"""
    if provider == "claude":
        api_key = os.environ.get('ANTHROPIC_API_KEY')
        if not api_key:
            click.echo("Error: Set ANTHROPIC_API_KEY", err=True)
            sys.exit(1)
    else:
        api_key = os.environ.get('GEMINI_API_KEY')
        if not api_key and not SERVICE_ACCOUNT_JSON.exists():
            click.echo("Error: Set GEMINI_API_KEY or place service account JSON", err=True)
            sys.exit(1)
        if SERVICE_ACCOUNT_JSON.exists():
            click.echo(f"Using service account: {SERVICE_ACCOUNT_JSON.name}")

    pdf_path = Path(pdf_path)
    data, confidence = process_pdf(pdf_path, api_key, provider=provider, model=model,
                                   use_fewshot=not no_fewshot)

    click.echo("\n--- Extracted Data ---")
    for key, value in data.items():
        if not key.startswith('_'):
            conf = confidence.get(key)
            conf_str = f" (confidence: {conf:.0%})" if conf is not None else ""
            click.echo(f"  {key}: {value}{conf_str}")


@cli.command()
@click.argument('input_dir', type=click.Path(exists=True))
def analyze(input_dir: str):
    """Analyze a directory and show PDF summary (no API needed)"""
    input_dir = Path(input_dir)
    pdf_files = list(input_dir.glob('**/*.pdf'))

    click.echo(f"\nDirectory: {input_dir}")
    click.echo(f"PDF files found: {len(pdf_files)}")

    excel_files = list(input_dir.glob('**/*.xlsx'))
    click.echo(f"Excel files found: {len(excel_files)}")

    # Check examples
    examples = load_examples()
    click.echo(f"Few-shot examples: {len(examples)}")

    if pdf_files:
        click.echo("\nPDF files:")
        for pdf in pdf_files[:10]:
            result = subprocess.run(
                ["pdftotext", str(pdf), "-"],
                capture_output=True, text=True
            )
            text_len = len(result.stdout.strip())
            text_status = "text" if text_len > 100 else "image"
            click.echo(f"  {pdf.name} ({text_status})")

        if len(pdf_files) > 10:
            click.echo(f"  ... and {len(pdf_files) - 10} more")


@cli.command()
def examples():
    """List available few-shot examples"""
    exs = load_examples()

    if not exs:
        click.echo("No examples found in examples/ directory")
        click.echo("\nTo add examples:")
        click.echo("  1. pdftoppm -png -r 150 -singlefile your.pdf examples/example_name")
        click.echo("  2. Create examples/example_name.json with correct values")
        return

    click.echo(f"Found {len(exs)} examples:\n")
    for ex in exs:
        click.echo(f"  {ex['name']}:")
        for k, v in ex['output'].items():
            if v is not None:
                click.echo(f"    {k}: {v}")
        click.echo()


if __name__ == "__main__":
    cli()
